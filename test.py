from codecs import unicode_escape_decode
from encodings import utf_8
import subprocess
import argparse
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import os
import time 
import pandas as pd
DEFAULT_EXEC = "/data1/publicUser/wjh/richman/build/richman "
EXPIRE_TIME_MS = 1000
TESTDIR = 'testcase/testcase_g2'
GREEN = PatternFill(start_color='7FFF00',
                    end_color='7FFF00', fill_type='solid')
RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
def readfilepair(args):
    dirname= os.listdir(args.dir) #得到文件夹下的所有文件名称
    #dirname = ['pay rent']
    files = []
    for path in dirname:
        dir = os.path.join(args.dir,path)
        file = os.listdir(dir)
        #file = ["NO1-9.in","NO1-9.out"]
        file.sort()
        for i in file:
            files.append(os.path.join(dir,i))
    file_pairs = []
    for i in range(int(len(files)/2)):
        file_pairs.append((files[2*i],files[2*i+1]))
    return file_pairs

def textprocess(text):
    #text = "10000\n1234\n"+"step 6\ny\nprint\n"+'quit\n'
    tx = text.splitlines()
    text = ''
    for i in tx:
        if i[:10].lower() == 'preset seq':
            order = i[11:]
            d = {"Q":"1","A":"2","S":"3","J":"4","q":"1","a":"2","s":"3","j":"4"}
            numorder = ''
            for i in order:
                numorder += d[i]
            continue
        if i[:6].lower() == 'preset':
            text = text+ '\n' + i 
        else:
            text = text +'\n\n' + i
    text = '10000\n'+numorder+text.replace("\r\n",'\n')+'\n\nprint'+'\nquit\n'
    return text

def export_to_xlsx(file_name, res):
    wb = Workbook()
    ws = wb.active  # work sheet

    row = 0
    res_keys = list(res.keys())
    res_keys.sort()
    row = row + 1
    ws.cell(row, 1).value = 'testname'
    ws.cell(row, 2).value = 'PASS/FAIL'
    #ws.cell(row, 3).value = '%d/%d' % (cnt, len(res_keys))
    ws.cell(row, 3).value = 'runtime'
    ws.cell(row,4).value = 'error'
    alltime = 0
    cnt = 0
    for test in res_keys:   
        if res[test][1]:
            cnt = cnt + 1  # count True
        row = row + 1
        ws.cell(row, 1).value = test
        ws.cell(row, 2).value = "PASS" if res[test][1] else 'FAIL'
        ws.cell(row, 2).fill = GREEN if res[test][1] else RED
        alltime += res[test][0]
        ws.cell(row, 3).value = round(res[test][0],3)
        ws.cell(row, 4).value = str(res[test][2:])

    row = row + 1
    ws.cell(row, 1).value = "alltest"
    ws.cell(row, 2).value = "{}/{}".format(cnt,len(res_keys))
    ws.cell(row, 3).value = round(alltime,3)
    #     for item in res[test]:
    #         row = row + 1
    #         ws.cell(row, 1).value = item[0]
    #         ws.cell(row, 2).value = item[1]
    #         ws.cell(row, 3).value = 'PASS' if item[2] else 'FAIL'
    #         ws.cell(row, 3).fill = GREEN if item[2] else RED
    #         ws.cell(row, 4).value = item[3]
    #         if item[3] >= EXPIRE_TIME_MS:
    #             ws.cell(row, 4).fill = RED
    #         pass # end for

    #     row = row + 1  # empty line
    #     pass

    wb.save(file_name)
    df=pd.read_excel(file_name)
    print(df)
    print("{}/{}".format(cnt,len(res_keys)))
    return

if __name__ == "__main__":
    args = argparse.ArgumentParser()
    args.add_argument('-d', '--dir', help='testcase dir',
                    default=TESTDIR, type=str)#改为测试在的dir
    args.add_argument('-n', '--name', help='process name',
                    default=DEFAULT_EXEC, type=str)#exe在的地方

    args = args.parse_args()

    file_pairs = readfilepair(args)#(file.in,file.out)

    res = {}

    succeed = 0
    fail = 0
    for files in file_pairs:
        fin = open(files[0],'rb')
        text = fin.read().decode()
        text = textprocess(text)
        res[files[0]] = []
        print("test in "+files[0]+":")
        t = time.time()
        try:
            child = subprocess.Popen(args.name, universal_newlines=True,shell =True,
                stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out = child.communicate(text,timeout=EXPIRE_TIME_MS/1000)
            out = out[0].splitlines()
        except subprocess.TimeoutExpired:

            #print("time out in {}".format(files[0]))
            pass
        #print(out[0].splitlines())
        t = time.time() - t
        res[files[0]].append(t)
        fout = open(files[1],'rb')
        ans = fout.read().decode()
        ans = ans.replace('\n\n','\n').splitlines()
        ans = [i.strip()  for i in ans if i]
        ans = [i.strip()  for i in ans if i]
        out = [i.strip() for i in out]
        lenans = len(ans)
        out = out[-lenans:]
        out[0] = out[0][-len(ans[0]):]
        ans.sort()
        out.sort()
        count = 0
        errorline = 0
        error = {}
        for i in range(lenans):#倒着比
            outlist = []
            anslist = []
            try:
                if ans[-i-1]== out[-i-1]:
                    count += 1
                else:
                    errorline +=1
                    error["ans{}".format(errorline)] = ans[-i-1]
                    error["out{}".format(errorline)] = out[-i-1]
                    #print(text)
                    print(error)
            except:
                pass
        if count == lenans:
            res[files[0]].append(1)
            print('succeed')
            succeed+=1
        else:
            res[files[0]].append(0)
            print("fail")
            fail+=1
        res[files[0]].append(error)
        #print(out[0])
    export_to_xlsx('./result.xlsx',res)